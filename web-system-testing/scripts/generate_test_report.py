"""
生成标准化测试报告
报告格式严格按照参考模板，包含四个Sheet：
1. 测试概要
2. 详细测试结果（每个操作一行，13列）
3. Bug列表（9列，含用例ID）
4. 测试结果统计

【重要】每个功能点的增、查、改、删操作必须分行记录
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import json


# 操作类型底色配置
OPERATION_COLOR_MAP = {
    "新增": "90EE90",   # 浅绿色
    "查看": "ADD8E6",   # 浅蓝色
    "编辑": "FFFACD",   # 浅黄色
    "删除": "FFB6C1",   # 浅红色
    "审核": "FFDAB9",   # 浅橙色
    "不适用": "D3D3D3"  # 浅灰色
}


def create_test_report(
    test_run_name: str,
    test_module: str,
    test_time: str,
    test_type: str,
    total_functions: int,
    test_results: list,
    bug_list: list,
    blocked_items: list,
    output_dir: str
):
    """
    创建标准化测试报告

    Args:
        test_run_name: 测试轮次名称，如"第十二轮"
        test_module: 测试模块名称，如"采伐证核发"
        test_time: 测试时间，格式YYYY-MM-DD
        test_type: 测试类型，如"功能全面测试"
        total_functions: 功能点总数（不同功能点的数量）
        test_results: 详细测试结果列表（每行一条记录，一个功能的一个操作）
        bug_list: Bug列表
        blocked_items: 阻塞功能点列表
        output_dir: 输出目录
    """
    wb = Workbook()

    # 计算统计数据
    tested_count = len([r for r in test_results if r.get("实际结果")])
    passed_count = len([r for r in test_results if r.get("实际结果") and r.get("备注") != "失败"])
    failed_count = len([r for r in test_results if r.get("备注") == "失败"])
    blocked_count = len([r for r in test_results if r.get("备注") == "阻塞"])
    pass_rate = f"{(passed_count / tested_count * 100):.1f}%" if tested_count > 0 else "0%"

    # === Sheet 1: 测试概要 ===
    ws_summary = wb.active
    ws_summary.title = "测试概要"

    summary_data = [
        [f"{test_run_name}测试报告 - {test_module}", None, None, None],
        [None, None, None, None],
        ["测试时间", test_time, None, None],
        ["测试模块", test_module, None, None],
        ["测试类型", test_type, None, None],
        ["功能点总数", total_functions, None, None],
        [None, None, None, None],
        ["测试结果统计", None, None, None],
        ["总功能点数", total_functions, None, None],
        ["已测试", tested_count, None, None],
        ["通过", passed_count, None, None],
        ["失败", failed_count, None, None],
        ["阻塞", blocked_count, None, None],
        ["通过率", pass_rate, None, None],
    ]

    for row in summary_data:
        ws_summary.append(row)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 55

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    ws_summary['A1'].font = title_font
    ws_summary['A8'].font = header_font

    # 添加阻塞功能点说明
    if blocked_items:
        ws_summary.append(["阻塞功能点", None, None, None])
        for item in blocked_items:
            ws_summary.append([item, None, None, None])

    # === Sheet 2: 详细测试结果 ===
    ws_detail = wb.create_sheet("详细测试结果")

    # 13列格式（含用例ID、UI测试、性能测试、截图名称）
    detail_headers = [
        "序号", "用例ID", "功能点", "操作类型", "优先级",
        "操作步骤", "预期结果", "实际结果", "测试时间",
        "UI测试", "性能测试", "截图名称", "备注"
    ]

    ws_detail.append(detail_headers)

    for result in test_results:
        row = [
            result.get("序号", ""),
            result.get("用例ID", ""),
            result.get("功能点", ""),
            result.get("操作类型", ""),
            result.get("优先级", "P1"),
            result.get("操作步骤", ""),
            result.get("预期结果", ""),
            result.get("实际结果", ""),
            result.get("测试时间", ""),
            result.get("UI测试", ""),
            result.get("性能测试", ""),
            result.get("截图名称", ""),
            result.get("备注", ""),
        ]
        ws_detail.append(row)

    # 设置列宽
    col_widths = [6, 10, 18, 10, 8, 55, 40, 40, 18, 10, 12, 30, 20]
    for i, width in enumerate(col_widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = width

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_style = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font_style
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws_detail.iter_rows(min_row=1, max_row=len(test_results)+1, min_col=1, max_col=13):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 设置操作类型单元格的底色
    for row_num in range(2, len(test_results) + 2):
        operation_cell = ws_detail.cell(row=row_num, column=4)  # 操作类型在第4列
        operation = operation_cell.value
        if operation in OPERATION_COLOR_MAP:
            color = OPERATION_COLOR_MAP[operation]
            operation_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # === Sheet 3: Bug列表 ===
    ws_bug = wb.create_sheet("Bug列表")

    # 9列格式（含用例ID）
    bug_headers = ["Bug ID", "用例ID", "功能点", "严重程度", "问题描述", "预期结果", "实际结果", "状态", "建议"]
    ws_bug.append(bug_headers)

    for i, bug in enumerate(bug_list, 1):
        row = [
            bug.get("Bug ID", f"BUG_{i:03d}"),
            bug.get("用例ID", ""),
            bug.get("功能点", ""),
            bug.get("严重程度", "一般"),
            bug.get("问题描述", ""),
            bug.get("预期结果", ""),
            bug.get("实际结果", ""),
            bug.get("状态", "待确认"),
            bug.get("建议", ""),
        ]
        ws_bug.append(row)

    ws_bug.column_dimensions['A'].width = 10
    ws_bug.column_dimensions['B'].width = 10
    ws_bug.column_dimensions['C'].width = 18
    ws_bug.column_dimensions['D'].width = 10
    ws_bug.column_dimensions['E'].width = 50
    ws_bug.column_dimensions['F'].width = 40
    ws_bug.column_dimensions['G'].width = 35
    ws_bug.column_dimensions['H'].width = 12
    ws_bug.column_dimensions['I'].width = 35

    for cell in ws_bug[1]:
        cell.fill = header_fill
        cell.font = header_font_style
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws_bug.iter_rows(min_row=1, max_row=len(bug_list)+1, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 设置严重程度颜色
    severity_colors = {"阻塞": "FF0000", "严重": "FFC7CE", "一般": "FFEB9C", "低": "C6EFCE"}
    for row_num in range(2, len(bug_list)+2):
        severity_cell = ws_bug.cell(row=row_num, column=4)
        if severity_cell.value in severity_colors:
            color = severity_colors[severity_cell.value]
            severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # === Sheet 4: 测试结果统计 ===
    ws_stats = wb.create_sheet("测试结果统计")

    stats_data = [
        ["测试结果统计", None, None, None],
        [None, None, None, None],
        ["功能点", "数量", "占比", None],
        ["总功能点数", total_functions, "100%", None],
        ["已测试", tested_count, f"{(tested_count/total_functions*100):.1f}%" if total_functions > 0 else "0%", None],
        ["通过", passed_count, f"{(passed_count/total_functions*100):.1f}%" if total_functions > 0 else "0%", None],
        ["失败", failed_count, f"{(failed_count/total_functions*100):.1f}%" if total_functions > 0 else "0%", None],
        ["阻塞", blocked_count, f"{(blocked_count/total_functions*100):.1f}%" if total_functions > 0 else "0%", None],
        [None, None, None, None],
        ["优先级分布", None, None, None],
        ["优先级", "功能点数", "已测试", "通过"],
    ]

    # 计算优先级分布
    priority_stats = {}
    for r in test_results:
        p = r.get("优先级", "P1")
        if p not in priority_stats:
            priority_stats[p] = {"total": 0, "tested": 0, "passed": 0}
        priority_stats[p]["total"] += 1
        if r.get("实际结果"):
            priority_stats[p]["tested"] += 1
        if r.get("备注") != "失败":
            priority_stats[p]["passed"] += 1

    for p in sorted(priority_stats.keys()):
        stats_data.append([
            p,
            priority_stats[p]["total"],
            priority_stats[p]["tested"],
            priority_stats[p]["passed"]
        ])

    for row in stats_data:
        ws_stats.append(row)

    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 15
    ws_stats.column_dimensions['C'].width = 15
    ws_stats.column_dimensions['D'].width = 15

    ws_stats['A1'].font = title_font
    for row in ws_stats.iter_rows(min_row=10, max_row=10, min_col=1, max_col=1):
        for cell in row:
            cell.font = header_font

    # 保存文件
    output_path = Path(output_dir) / f"{test_run_name}测试详细报告_{test_module}_{test_time.replace('-', '')}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return str(output_path)


def load_test_results_from_json(json_path: str) -> dict:
    """从JSON文件加载测试结果"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


if __name__ == "__main__":
    # 测试数据 - 多行记录示例
    test_results = [
        {
            "序号": 1,
            "用例ID": "TC-001",
            "功能点": "一般林木申请",
            "操作类型": "新增",
            "优先级": "P0",
            "操作步骤": "1. 点击一般林木申请Tab 2. 点击新增按钮 3. 填写申请人姓名、证件号码等字段 4. 点击保存按钮",
            "预期结果": "弹出保存成功提示，列表显示新创建的记录",
            "实际结果": "弹出保存成功提示，列表显示新创建的记录",
            "测试时间": "2026-04-22 10:30",
            "UI测试": "无",
            "性能测试": "1250ms",
            "截图名称": "TC-001_一般林木申请_新增.png",
            "备注": ""
        },
        {
            "序号": 2,
            "用例ID": "TC-002",
            "功能点": "一般林木申请",
            "操作类型": "查看",
            "优先级": "P0",
            "操作步骤": "1. 在一般林木申请列表中点击任意记录查看详情",
            "预期结果": "显示完整的采伐申请表单信息",
            "实际结果": "显示完整的采伐申请表单信息",
            "测试时间": "2026-04-22 10:35",
            "UI测试": "无",
            "性能测试": "890ms",
            "截图名称": "TC-002_一般林木申请_查看.png",
            "备注": ""
        },
        {
            "序号": 3,
            "用例ID": "TC-003",
            "功能点": "一般林木申请",
            "操作类型": "编辑",
            "优先级": "P0",
            "操作步骤": "1. 点击编辑按钮 2. 修改联系电话字段 3. 点击保存按钮",
            "预期结果": "记录更新成功，列表显示修改后的信息",
            "实际结果": "记录更新成功，列表显示修改后的信息",
            "测试时间": "2026-04-22 10:40",
            "UI测试": "无",
            "性能测试": "1100ms",
            "截图名称": "TC-003_一般林木申请_编辑.png",
            "备注": ""
        },
        {
            "序号": 4,
            "用例ID": "TC-004",
            "功能点": "待审核",
            "操作类型": "查看",
            "优先级": "P1",
            "操作步骤": "1. 点击待审核Tab 2. 查看待审核列表",
            "预期结果": "显示待审核记录列表，每条记录包含申请人、市县、采伐蓄积等",
            "实际结果": "显示待审核记录列表，包含申请人、市县、采伐蓄积等字段",
            "测试时间": "2026-04-22 10:50",
            "UI测试": "无",
            "性能测试": "1500ms",
            "截图名称": "TC-004_待审核_查看.png",
            "备注": ""
        },
    ]

    bug_list = []
    blocked_items = []

    output = create_test_report(
        test_run_name="第十二轮",
        test_module="采伐证核发",
        test_time="2026-04-22",
        test_type="功能全面测试",
        total_functions=2,
        test_results=test_results,
        bug_list=bug_list,
        blocked_items=blocked_items,
        output_dir=r"c:\PythonWork\autoTest\test_runs\testRun-012-20260422"
    )

    print(f"测试报告已生成: {output}")
